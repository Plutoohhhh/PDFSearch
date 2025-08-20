import sys

import pandas as pd
from PySide6.QtCore import QStandardPaths
from PySide6.QtWidgets import QWidget, QApplication, QFileDialog, QComboBox, QHBoxLayout, QLabel
from openpyxl import load_workbook

from model.ExcelModel import ExcelModel
from ui.TagsWidgetUI import Ui_TagsWidget


class TagsWidget(QWidget, Ui_TagsWidget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.resetUi()
        self.bind()
        # 初始化模型
        self.model = ExcelModel([])

    def bind(self):
        self.btn_selectExcel.clicked.connect(self.on_btn_selectExcel)
        # 初始化选中
        self.category_combo.currentIndexChanged.connect(self.update_model_apn)

    def resetUi(self):
        self.category_combo = QComboBox()
        self.model_combo = QComboBox()
        self.apn_combo = QComboBox()
        # 添加每一行
        self.add_filter_row("Material Category", self.category_combo)
        self.add_filter_row("Material Model", self.model_combo)
        self.add_filter_row("APN", self.apn_combo)
        self.label_status.hide()

    def add_filter_row(self, label_text, combo_box, items=None):
        items = items or []
        combo_box.addItems(items)
        row_layout = QHBoxLayout()
        label = QLabel(label_text)
        label.setFixedWidth(180)
        row_layout.addWidget(label)
        row_layout.addWidget(combo_box)
        self.verticalLayout.addLayout(row_layout)

    def update_model_apn(self):
        index = self.category_combo.currentIndex()
        if index < 0:
            return

        self.model_combo.clear()
        self.apn_combo.clear()

        if index < len(self.model.grouped_unique_second_col):
            self.model_combo.addItems(self.model.grouped_unique_second_col[index])

        if index < len(self.model.grouped_unique_third_col):
            self.apn_combo.addItems(self.model.grouped_unique_third_col[index])


    def on_btn_selectExcel(self):
        # 获取桌面路径
        desktop_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
        # 用户选择fleet data Excel文件的目录路径
        excel_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", desktop_path,
                                                    "Excel files (*.xlsx *.xls *.csv)")
        if excel_path:
            self.lineEdit_excel_path.setText(excel_path)
            self.excel_read(excel_path)

    # 读取excel有效数据，做成过滤标签
    def excel_read(self, excel_path):
        # # 加载工作簿和工作表
        # workbook_excel = load_workbook(excel_path, read_only=True)
        # sheet = workbook_excel.active
        # for row in sheet.iter_rows():
        #     for cell in row:
        #         print(cell.value)

        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            print("读取失败:", e)
            return

        # ✅ 指定需要保留的列名
        desired_columns = ['Material Categroy', 'Material Model', 'APN']

        # 筛选出存在的列（防止列名不存在时报错）
        existing_columns = [col for col in desired_columns if col in df.columns]

        # ✅ 使用 pandas 前向填充（ffill）处理 NaN 值
        df = df[existing_columns].ffill()

        # 🌟 将数据从 DataFrame 转为列优先格式（每一列是一个列表）
        headers = existing_columns
        data_by_column = [df[col].astype(str).tolist() for col in existing_columns]

        # 创建并设置模型
        self.model = ExcelModel(data_by_column=data_by_column, headers=headers, parent=self)
        self.tableView_excel.setModel(self.model)

        # 自动调整列宽
        self.tableView_excel.resizeColumnsToContents()

        # 打印调试信息
        print("筛选后的数据：", self.model._data)

        # 更新过滤器
        self.category_combo.clear()
        self.category_combo.addItems(self.model.grouped_unique_first_col)

        self.update_model_apn()





if __name__ == "__main__":
    app = QApplication(sys.argv)
    viewer = TagsWidget()
    viewer.show()
    sys.exit(app.exec())